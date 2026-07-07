#!/usr/bin/env python3
"""
优惠券批量生成器 - Coupon Batch Generator
==========================================
本地 Python HTTP 服务器 + 浏览器前端，将 Excel 券码与模板图片合成为
带独立二维码和券码文字的多页 PDF，用于批量印刷优惠券。

用法: python3 coupon_web.py
"""

# Windows 默认 stdout 编码是 cp1252，无法输出中文；强制 UTF-8
import sys
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr and hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import http.server
import socketserver
import json
import os
import io
import sys
import csv
import re
import tempfile
import threading
import webbrowser
from urllib.parse import urlparse, parse_qs

# ==================== 第三方库 ====================
import qrcode
from PIL import Image, ImageDraw, ImageFont
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


# ==================== 配置 ====================
HOST = 'localhost'
PORT = 8888


# ==================== 全局状态（线程安全） ====================
state_lock = threading.Lock()
state = {
    'template_image': None,    # PIL Image 对象
    'template_bytes': None,    # PNG bytes（用于前端预览）
    'template_size': None,     # (width, height)
    'codes': [],               # 券码列表
    'codes_filename': None,
}


# ==================== 工具函数 ====================

def _find_chinese_font():
    """跨平台查找中文字体路径，返回可用的 .ttc/.ttf 文件。"""
    import platform
    candidates = []
    if platform.system() == 'Darwin':  # macOS
        candidates = [
            '/System/Library/Fonts/PingFang.ttc',
            '/System/Library/Fonts/STHeiti Medium.ttc',
            '/Library/Fonts/Arial Unicode.ttf',
        ]
    elif platform.system() == 'Windows':
        windir = os.environ.get('WINDIR', 'C:\\Windows')
        candidates = [
            os.path.join(windir, 'Fonts', 'msyh.ttc'),      # 微软雅黑
            os.path.join(windir, 'Fonts', 'msyh.ttf'),
            os.path.join(windir, 'Fonts', 'simhei.ttf'),    # 黑体
            os.path.join(windir, 'Fonts', 'simsun.ttc'),    # 宋体
            os.path.join(windir, 'Fonts', 'simfang.ttf'),   # 仿宋
        ]
    else:  # Linux 等
        candidates = [
            '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
            '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
            '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
            '/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf',
        ]
    for path in candidates:
        if os.path.isfile(path):
            return path
    return None


def generate_qr_image(code, size=200):
    """为给定券码生成 QR 码 PIL 图片。"""
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )
    qr.add_data(str(code))
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img = img.convert('RGB')
    if size != img.size[0]:
        img = img.resize((size, size), Image.LANCZOS)
    return img


def parse_excel_file(filepath, col=0, skip_header=False):
    """
    解析 Excel/CSV 文件，返回指定列的券码列表。

    Args:
        filepath: 文件路径
        col: 列索引 (0=A, 1=B, ...)
        skip_header: 是否跳过第一行
    Returns:
        list[str]: 券码列表
    """
    ext = os.path.splitext(filepath)[1].lower()
    codes = []

    if ext == '.csv':
        # UTF-8 with BOM 兼容
        for encoding in ('utf-8-sig', 'utf-8', 'gbk', 'gb2312'):
            try:
                with open(filepath, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                break
            except (UnicodeDecodeError, csv.Error):
                rows = None
        if rows is None:
            raise ValueError("CSV 文件编码无法识别，请使用 UTF-8 编码")

    elif ext == '.xlsx' and HAS_OPENPYXL:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        wb.close()

    elif ext in ('.xls', '.et') and HAS_XLRD:
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]

    elif ext == '.et':
        # 尝试 xlrd 读取 .et（WPS 表格部分格式兼容 xls）
        if HAS_XLRD:
            try:
                wb = xlrd.open_workbook(filepath)
                ws = wb.sheet_by_index(0)
                rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
            except Exception:
                raise ValueError(".et 文件解析失败，请转换为 .xlsx 或 .csv 格式")
        else:
            raise ValueError("无法解析 .et 文件，请安装 xlrd 库或转换为 .xlsx/.csv 格式")

    else:
        raise ValueError(
            f"不支持的文件格式: {ext}。支持 .xlsx, .xls, .et, .csv"
        )

    # 提取指定列的券码
    start = 1 if skip_header and len(rows) > 0 else 0
    for row in rows[start:]:
        if col < len(row) and row[col] is not None:
            code = str(row[col]).strip()
            if code:
                codes.append(code)

    return codes


def generate_pdf(codes, template_image, qr_config, text_config):
    """
    生成多页 PDF，每页 = 模板图片 + 对应券码的二维码 + 券码文字。

    坐标系：浏览器左上角原点，y 向下增加
    PDF 坐标系：左下角原点，y 向上增加
    转换：pdf_y = img_height - browser_y - element_height
    """
    # 确保模板为 RGB 模式（打印需要）
    if template_image.mode in ('RGBA', 'LA') or \
       (template_image.mode == 'P' and 'transparency' in template_image.info):
        background = Image.new('RGB', template_image.size, (255, 255, 255))
        if template_image.mode == 'P':
            template_image = template_image.convert('RGBA')
        mask = template_image.split()[-1] if template_image.mode in ('RGBA', 'LA') else None
        background.paste(template_image, mask=mask)
        template_image = background
    elif template_image.mode != 'RGB':
        template_image = template_image.convert('RGB')

    img_w, img_h = template_image.size
    template_reader = ImageReader(template_image)

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(img_w, img_h))

    qr_x = qr_config['x']
    qr_y = qr_config['y']
    qr_size = qr_config['size']

    text_x = text_config['x']
    text_y = text_config['y']
    text_font_size = text_config['font_size']
    text_show = text_config.get('show', True)

    for code in codes:
        # 1. 绘制模板底图
        c.drawImage(template_reader, 0, 0, width=img_w, height=img_h)

        # 2. 生成并绘制二维码
        qr_img = generate_qr_image(code, qr_size)
        qr_reader = ImageReader(qr_img)
        # 坐标转换：浏览器 y → PDF y
        pdf_qr_y = img_h - qr_y - qr_size
        c.drawImage(qr_reader, qr_x, pdf_qr_y, width=qr_size, height=qr_size)

        # 3. 绘制券码文字（可选、可选中复制）
        if text_show:
            # drawString 的 y 是基线位置
            pdf_text_y = img_h - text_y - text_font_size

            # 绘制半透明白色背景
            text_w = stringWidth(str(code), "Helvetica", text_font_size)
            padding = max(2, text_font_size * 0.1)
            c.saveState()
            c.setFillColorRGB(1, 1, 1)
            c.setFillAlpha(0.6)
            c.rect(
                text_x - padding,
                pdf_text_y - padding,
                text_w + 2 * padding,
                text_font_size + 2 * padding,
                fill=1, stroke=0
            )
            c.restoreState()

            # 绘制文字（真实文字，非图片像素，可选中复制）
            c.setFillColorRGB(0, 0, 0)
            c.setFont("Helvetica", text_font_size)
            c.drawString(text_x, pdf_text_y, str(code))

        c.showPage()

    c.save()
    buf.seek(0)
    return buf


def parse_multipart(body, boundary):
    """解析 multipart/form-data，返回 {field_name: {'content': bytes, 'filename': str|None}}。"""
    boundary_bytes = boundary.encode('utf-8')
    parts = body.split(b'--' + boundary_bytes)
    files = {}

    for part in parts:
        if not part or part.strip() in (b'', b'--', b'--\r\n'):
            continue
        header_end = part.find(b'\r\n\r\n')
        if header_end == -1:
            continue
        headers_raw = part[:header_end].decode('utf-8', errors='ignore')
        content = part[header_end + 4:]
        if content.endswith(b'\r\n'):
            content = content[:-2]

        name_match = re.search(r'name="([^"]*)"', headers_raw)
        filename_match = re.search(r'filename="([^"]*)"', headers_raw)
        name = name_match.group(1) if name_match else ''
        filename = filename_match.group(1) if filename_match else None
        files[name] = {'content': content, 'filename': filename}

    return files


# ==================== HTML 前端模板 ====================
HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>优惠券批量生成器</title>
<style>
/* ==================== 主题变量 ==================== */
:root {
  --bg: #f0f2f5;
  --bg-card: #ffffff;
  --bg-input: #f5f6f8;
  --bg-hover: #eef1f6;
  --text: #1a1d28;
  --text-muted: #6b7280;
  --text-light: #9ca3af;
  --border: #e2e8f0;
  --border-strong: #cbd5e1;
  --accent: #4a9eff;
  --accent-hover: #3a8eef;
  --accent-light: rgba(74, 158, 255, 0.08);
  --danger: #ef4444;
  --danger-hover: #dc2626;
  --success: #10b981;
  --warning: #f59e0b;
  --shadow-sm: 0 1px 2px rgba(0,0,0,0.04);
  --shadow: 0 1px 3px rgba(0,0,0,0.06), 0 1px 2px rgba(0,0,0,0.04);
  --shadow-lg: 0 10px 30px rgba(0,0,0,0.08), 0 5px 15px rgba(0,0,0,0.04);
  --radius: 14px;
  --radius-sm: 8px;
  --transition: 0.2s cubic-bezier(0.4, 0, 0.2, 1);
  --font-sans: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', 'PingFang SC', 'Microsoft YaHei', sans-serif;
  --font-mono: 'SF Mono', 'Fira Code', Menlo, Consolas, 'Liberation Mono', monospace;
}

[data-theme="dark"] {
  --bg: #0f1117;
  --bg-card: #1a1d28;
  --bg-input: #252836;
  --bg-hover: #2a2d3a;
  --text: #e8e9ed;
  --text-muted: #8b8d98;
  --text-light: #6b7280;
  --border: #2a2d3a;
  --border-strong: #3a3d4a;
  --accent: #4a9eff;
  --accent-hover: #5aaeff;
  --accent-light: rgba(74, 158, 255, 0.12);
  --shadow-sm: 0 1px 2px rgba(0,0,0,0.2);
  --shadow: 0 1px 3px rgba(0,0,0,0.3);
  --shadow-lg: 0 10px 30px rgba(0,0,0,0.4);
}

/* ==================== 基础重置 ==================== */
* { box-sizing: border-box; margin: 0; padding: 0; }

html, body { height: 100%; }

body {
  font-family: var(--font-sans);
  background: var(--bg);
  color: var(--text);
  line-height: 1.6;
  font-size: 14px;
  overflow: hidden;
  transition: background var(--transition), color var(--transition);
}

/* ==================== 顶部导航 ==================== */
.header {
  height: 60px;
  background: var(--bg-card);
  border-bottom: 1px solid var(--border);
  display: flex;
  align-items: center;
  justify-content: space-between;
  padding: 0 24px;
  box-shadow: var(--shadow-sm);
  z-index: 100;
  position: relative;
}

.header-left {
  display: flex;
  align-items: center;
  gap: 12px;
}

.header-logo {
  width: 32px;
  height: 32px;
  border-radius: 8px;
  background: linear-gradient(135deg, var(--accent), #7b5fff);
  display: flex;
  align-items: center;
  justify-content: center;
  font-size: 18px;
  color: #fff;
  font-weight: 700;
}

.header-title {
  font-size: 17px;
  font-weight: 700;
  letter-spacing: -0.02em;
}

.header-subtitle {
  font-size: 12px;
  color: var(--text-muted);
  margin-left: 4px;
}

.header-controls {
  display: flex;
  align-items: center;
  gap: 8px;
}

.icon-btn {
  width: 36px;
  height: 36px;
  border-radius: var(--radius-sm);
  border: 1px solid var(--border);
  background: var(--bg-card);
  color: var(--text);
  cursor: pointer;
  display: flex;
  align-items: center;
  justify-content: center;
  font-size: 14px;
  font-weight: 600;
  transition: all var(--transition);
}

.icon-btn:hover {
  background: var(--bg-hover);
  border-color: var(--border-strong);
  transform: translateY(-1px);
}

.icon-btn:active { transform: translateY(0); }

/* ==================== 主布局 ==================== */
.layout {
  display: flex;
  height: calc(100vh - 60px);
}

/* ==================== 左侧控制面板 ==================== */
.control-panel {
  width: 380px;
  min-width: 380px;
  background: var(--bg-card);
  border-right: 1px solid var(--border);
  overflow-y: auto;
  padding: 20px;
  display: flex;
  flex-direction: column;
  gap: 16px;
}

.control-panel::-webkit-scrollbar { width: 6px; }
.control-panel::-webkit-scrollbar-track { background: transparent; }
.control-panel::-webkit-scrollbar-thumb { background: var(--border-strong); border-radius: 3px; }

/* 步骤卡片 */
.step-card {
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 20px;
  position: relative;
  transition: border-color var(--transition), box-shadow var(--transition);
}

.step-card:hover { border-color: var(--border-strong); }

.step-card.active {
  border-color: var(--accent);
  box-shadow: 0 0 0 3px var(--accent-light);
}

.step-header {
  display: flex;
  align-items: center;
  gap: 10px;
  margin-bottom: 16px;
}

.step-number {
  width: 28px;
  height: 28px;
  border-radius: 8px;
  background: var(--accent-light);
  color: var(--accent);
  display: flex;
  align-items: center;
  justify-content: center;
  font-family: var(--font-mono);
  font-size: 13px;
  font-weight: 700;
  flex-shrink: 0;
}

.step-title {
  font-size: 15px;
  font-weight: 700;
}

.step-desc {
  font-size: 12px;
  color: var(--text-muted);
  margin-bottom: 14px;
}

/* 上传区域 */
.upload-zone {
  border: 2px dashed var(--border-strong);
  border-radius: var(--radius-sm);
  padding: 24px 16px;
  text-align: center;
  cursor: pointer;
  transition: all var(--transition);
  position: relative;
}

.upload-zone:hover {
  border-color: var(--accent);
  background: var(--accent-light);
}

.upload-zone.has-file {
  border-style: solid;
  border-color: var(--success);
  background: rgba(16, 185, 129, 0.04);
}

.upload-icon {
  font-size: 28px;
  margin-bottom: 8px;
  opacity: 0.6;
}

.upload-text {
  font-size: 13px;
  font-weight: 600;
  color: var(--text);
}

.upload-hint {
  font-size: 11px;
  color: var(--text-muted);
  margin-top: 4px;
}

.upload-zone input[type="file"] {
  position: absolute;
  inset: 0;
  opacity: 0;
  cursor: pointer;
}

/* 表单元素 */
.form-row {
  display: flex;
  align-items: center;
  gap: 10px;
  margin-top: 12px;
}

.form-row label {
  font-size: 13px;
  color: var(--text-muted);
  white-space: nowrap;
  display: flex;
  align-items: center;
  gap: 6px;
  cursor: pointer;
}

.form-row input[type="number"],
.form-row select {
  flex: 1;
  height: 34px;
  padding: 0 10px;
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  background: var(--bg-input);
  color: var(--text);
  font-size: 13px;
  font-family: var(--font-mono);
  transition: border-color var(--transition);
  outline: none;
  min-width: 0;
}

.form-row input[type="number"]:focus,
.form-row select:focus {
  border-color: var(--accent);
  box-shadow: 0 0 0 3px var(--accent-light);
}

.form-row input[type="checkbox"] {
  width: 16px;
  height: 16px;
  accent-color: var(--accent);
  cursor: pointer;
}

.form-grid {
  display: grid;
  grid-template-columns: 1fr 1fr 1fr;
  gap: 8px;
  margin-top: 12px;
}

.form-grid .form-item {
  display: flex;
  flex-direction: column;
  gap: 4px;
}

.form-grid .form-item label {
  font-size: 11px;
  color: var(--text-muted);
  font-weight: 600;
  text-transform: uppercase;
  letter-spacing: 0.04em;
}

.form-grid .form-item input {
  height: 32px;
  padding: 0 8px;
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  background: var(--bg-input);
  color: var(--text);
  font-size: 13px;
  font-family: var(--font-mono);
  outline: none;
  transition: border-color var(--transition);
  width: 100%;
}

.form-grid .form-item input:focus {
  border-color: var(--accent);
  box-shadow: 0 0 0 3px var(--accent-light);
}

/* 信息展示 */
.info-box {
  margin-top: 12px;
  padding: 10px 12px;
  background: var(--bg-input);
  border-radius: var(--radius-sm);
  font-size: 12px;
  color: var(--text-muted);
  display: none;
}

.info-box.show { display: block; }

.info-box .info-count {
  font-size: 15px;
  font-weight: 700;
  color: var(--accent);
  font-family: var(--font-mono);
}

.codes-preview {
  margin-top: 8px;
  max-height: 120px;
  overflow-y: auto;
  display: none;
}

.codes-preview.show { display: block; }

.codes-preview::-webkit-scrollbar { width: 4px; }
.codes-preview::-webkit-scrollbar-thumb { background: var(--border-strong); border-radius: 2px; }

.code-item {
  padding: 4px 8px;
  font-family: var(--font-mono);
  font-size: 12px;
  color: var(--text);
  background: var(--bg-input);
  border-radius: 4px;
  margin-bottom: 3px;
  border-left: 2px solid var(--accent);
}

/* 按钮 */
.btn {
  height: 38px;
  padding: 0 16px;
  border-radius: var(--radius-sm);
  border: 1px solid var(--border);
  background: var(--bg-card);
  color: var(--text);
  font-size: 13px;
  font-weight: 600;
  cursor: pointer;
  transition: all var(--transition);
  display: inline-flex;
  align-items: center;
  justify-content: center;
  gap: 6px;
}

.btn:hover {
  background: var(--bg-hover);
  border-color: var(--border-strong);
}

.btn-clear {
  width: 100%;
  margin-top: 12px;
  color: var(--danger);
  border-color: var(--border);
}

.btn-clear:hover {
  background: rgba(239, 68, 68, 0.06);
  border-color: var(--danger);
}

.btn-generate {
  width: 100%;
  height: 48px;
  border: none;
  border-radius: var(--radius);
  background: linear-gradient(135deg, var(--accent), #7b5fff);
  color: #fff;
  font-size: 15px;
  font-weight: 700;
  cursor: pointer;
  transition: all var(--transition);
  box-shadow: 0 4px 14px rgba(74, 158, 255, 0.3);
  letter-spacing: 0.02em;
}

.btn-generate:hover:not(:disabled) {
  transform: translateY(-1px);
  box-shadow: 0 6px 20px rgba(74, 158, 255, 0.4);
}

.btn-generate:active:not(:disabled) { transform: translateY(0); }

.btn-generate:disabled {
  opacity: 0.6;
  cursor: not-allowed;
}

/* 配置区分隔 */
.config-section {
  padding-top: 14px;
  margin-top: 14px;
  border-top: 1px solid var(--border);
}

.config-section-title {
  font-size: 13px;
  font-weight: 700;
  color: var(--text);
  margin-bottom: 4px;
  display: flex;
  align-items: center;
  gap: 6px;
}

.config-section-title .dot {
  width: 8px;
  height: 8px;
  border-radius: 50%;
}

.dot-qr { background: var(--danger); }
.dot-text { background: var(--accent); }

/* ==================== 右侧预览面板 ==================== */
.preview-panel {
  flex: 1;
  display: flex;
  align-items: center;
  justify-content: center;
  padding: 24px;
  overflow: auto;
  background: var(--bg);
  position: relative;
}

.preview-container {
  position: relative;
  display: inline-block;
  box-shadow: var(--shadow-lg);
  border-radius: 4px;
  overflow: visible;
}

.preview-container.dragging * { user-select: none !important; }

.template-img {
  display: block;
  border-radius: 4px;
  max-width: 100%;
  max-height: 100%;
}

.preview-placeholder {
  text-align: center;
  color: var(--text-muted);
  padding: 60px 40px;
}

.preview-placeholder .ph-icon {
  font-size: 48px;
  opacity: 0.3;
  margin-bottom: 16px;
}

.preview-placeholder .ph-text {
  font-size: 15px;
  font-weight: 500;
}

.preview-placeholder .ph-hint {
  font-size: 13px;
  margin-top: 8px;
  opacity: 0.7;
}

/* 可拖拽元素 */
.draggable {
  position: absolute;
  cursor: grab;
  user-select: none;
  transition: box-shadow var(--transition), border-color var(--transition);
}

.draggable:active { cursor: grabbing; }

.draggable.selected {
  box-shadow: 0 0 0 2px var(--accent);
  z-index: 10;
}

/* 二维码元素 */
.qr-element {
  border: 2px dashed transparent;
  border-radius: 2px;
}

.qr-element:hover,
.qr-element.selected {
  border-color: var(--danger);
}

.qr-element img {
  width: 100%;
  height: 100%;
  display: block;
  pointer-events: none;
}

/* 券码文字元素 */
.text-element {
  background: rgba(255, 255, 255, 0.6);
  padding: 2px 6px;
  border-radius: 3px;
  font-family: var(--font-mono);
  color: #000;
  white-space: nowrap;
  line-height: 1.2;
  border: 2px dashed transparent;
}

.text-element:hover,
.text-element.selected {
  border-color: var(--accent);
}

/* 调整大小手柄 */
.resize-handle {
  position: absolute;
  bottom: -7px;
  right: -7px;
  width: 14px;
  height: 14px;
  cursor: nwse-resize;
  border-radius: 3px;
  border: 2px solid #fff;
  box-shadow: 0 1px 4px rgba(0,0,0,0.3);
  z-index: 11;
}

.resize-handle-qr { background: #ff6b6b; }
.resize-handle-text { background: #4a9eff; }

/* 拖拽提示 */
.drag-hint {
  position: absolute;
  bottom: 16px;
  left: 50%;
  transform: translateX(-50%);
  padding: 6px 14px;
  background: rgba(0, 0, 0, 0.7);
  color: #fff;
  font-size: 12px;
  border-radius: 20px;
  pointer-events: none;
  opacity: 0;
  transition: opacity var(--transition);
  white-space: nowrap;
  z-index: 50;
}

.drag-hint.show { opacity: 1; }

/* ==================== Toast 通知 ==================== */
.toast-container {
  position: fixed;
  bottom: 24px;
  left: 50%;
  transform: translateX(-50%);
  z-index: 10000;
  display: flex;
  flex-direction: column;
  gap: 8px;
  align-items: center;
  pointer-events: none;
}

.toast {
  padding: 12px 20px;
  border-radius: var(--radius-sm);
  font-size: 13px;
  font-weight: 600;
  color: #fff;
  box-shadow: var(--shadow-lg);
  opacity: 0;
  transform: translateY(20px);
  transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
  pointer-events: auto;
  max-width: 400px;
  display: flex;
  align-items: center;
  gap: 8px;
}

.toast.show {
  opacity: 1;
  transform: translateY(0);
}

.toast-success { background: var(--success); }
.toast-error { background: var(--danger); }
.toast-info { background: var(--accent); }

/* ==================== 加载遮罩 ==================== */
.loading-overlay {
  position: fixed;
  inset: 0;
  background: rgba(0, 0, 0, 0.5);
  -webkit-backdrop-filter: blur(4px);
  backdrop-filter: blur(4px);
  display: none;
  align-items: center;
  justify-content: center;
  flex-direction: column;
  gap: 20px;
  z-index: 9999;
}

.loading-overlay.show { display: flex; }

.spinner {
  width: 48px;
  height: 48px;
  border: 4px solid rgba(255, 255, 255, 0.2);
  border-top-color: #fff;
  border-radius: 50%;
  animation: spin 0.8s linear infinite;
}

@keyframes spin { to { transform: rotate(360deg); } }

.loading-text {
  color: #fff;
  font-size: 15px;
  font-weight: 600;
}

/* ==================== 工具类 ==================== */
.hidden { display: none !important; }

/* ==================== 响应式 ==================== */
@media (max-width: 960px) {
  .layout { flex-direction: column; height: auto; }
  .control-panel { width: 100%; min-width: 0; max-height: none; border-right: none; border-bottom: 1px solid var(--border); }
  .preview-panel { min-height: 400px; }
  body { overflow: auto; }
}
</style>
</head>
<body>

<!-- ==================== 顶部导航 ==================== -->
<header class="header">
  <div class="header-left">
    <div class="header-logo">C</div>
    <div>
      <span class="header-title" data-i18n="appTitle">优惠券批量生成器</span>
      <span class="header-subtitle" data-i18n="appSubtitle">批量生成带独立二维码的优惠券 PDF</span>
    </div>
  </div>
  <div class="header-controls">
    <button class="icon-btn" id="theme-toggle" title="Theme">🌙</button>
    <button class="icon-btn" id="lang-toggle" title="Language">EN</button>
  </div>
</header>

<!-- ==================== 主布局 ==================== -->
<div class="layout">
  <!-- 左侧控制面板 -->
  <aside class="control-panel">

    <!-- 步骤 1: 导入券码 -->
    <div class="step-card" id="step1">
      <div class="step-header">
        <div class="step-number">01</div>
        <div class="step-title" data-i18n="step1Title">导入券码数据</div>
      </div>
      <div class="step-desc" data-i18n="step1Desc">上传 Excel / CSV 文件，自动读取券码</div>

      <div class="upload-zone" id="excel-zone">
        <input type="file" id="excel-input" accept=".xlsx,.xls,.et,.csv">
        <div class="upload-icon">📄</div>
        <div class="upload-text" data-i18n="uploadExcel">选择 Excel 文件</div>
        <div class="upload-hint" data-i18n="uploadExcelHint">支持 .xlsx .xls .et .csv</div>
      </div>

      <div class="form-row">
        <label data-i18n="columnLabel">券码所在列</label>
        <select id="column-select">
          <option value="0">A</option>
          <option value="1">B</option>
          <option value="2">C</option>
          <option value="3">D</option>
          <option value="4">E</option>
        </select>
      </div>

      <div class="form-row">
        <label>
          <input type="checkbox" id="skip-header" checked>
          <span data-i18n="skipHeader">跳过第一行（标题行）</span>
        </label>
      </div>

      <div class="info-box" id="codes-info">
        <span data-i18n="codesCount">已读取</span>
        <span class="info-count" id="codes-count">0</span>
        <span data-i18n="codesUnit">个券码</span>
      </div>

      <div class="codes-preview" id="codes-preview"></div>

      <button class="btn btn-clear" id="clear-excel" data-i18n="clear">清空</button>
    </div>

    <!-- 步骤 2: 加载模板图片 -->
    <div class="step-card" id="step2">
      <div class="step-header">
        <div class="step-number">02</div>
        <div class="step-title" data-i18n="step2Title">加载模板图片</div>
      </div>
      <div class="step-desc" data-i18n="step2Desc">上传优惠券底图（含 logo、固定文案）</div>

      <div class="upload-zone" id="image-zone">
        <input type="file" id="image-input" accept="image/png,image/jpeg,image/bmp,.png,.jpg,.jpeg,.bmp">
        <div class="upload-icon">🖼️</div>
        <div class="upload-text" data-i18n="uploadImage">选择模板图片</div>
        <div class="upload-hint" data-i18n="uploadImageHint">支持 PNG / JPG / JPEG / BMP</div>
      </div>

      <div class="info-box" id="image-info">
        <span data-i18n="imageSize">图片尺寸：</span>
        <span class="info-count" id="image-size-text">—</span>
      </div>

      <button class="btn btn-clear" id="clear-image" data-i18n="clear">清空</button>
    </div>

    <!-- 步骤 3: 排版定位 -->
    <div class="step-card" id="step3">
      <div class="step-header">
        <div class="step-number">03</div>
        <div class="step-title" data-i18n="step3Title">排版定位</div>
      </div>
      <div class="step-desc" data-i18n="step3Desc">拖拽调整二维码和文字位置，所见即所得</div>

      <!-- 二维码配置 -->
      <div class="config-section">
        <div class="config-section-title">
          <span class="dot dot-qr"></span>
          <span data-i18n="qrPosition">二维码</span>
        </div>
        <div class="form-grid">
          <div class="form-item">
            <label>X</label>
            <input type="number" id="qr-x" value="50" min="0">
          </div>
          <div class="form-item">
            <label>Y</label>
            <input type="number" id="qr-y" value="50" min="0">
          </div>
          <div class="form-item">
            <label data-i18n="size">大小</label>
            <input type="number" id="qr-size" value="100" min="20" max="500">
          </div>
        </div>
      </div>

      <!-- 券码文字配置 -->
      <div class="config-section">
        <div class="config-section-title">
          <span class="dot dot-text"></span>
          <span data-i18n="textPosition">券码文字</span>
        </div>
        <div class="form-grid">
          <div class="form-item">
            <label>X</label>
            <input type="number" id="text-x" value="50" min="0">
          </div>
          <div class="form-item">
            <label>Y</label>
            <input type="number" id="text-y" value="200" min="0">
          </div>
          <div class="form-item">
            <label data-i18n="fontSize">字号</label>
            <input type="number" id="text-size" value="24" min="8" max="100">
          </div>
        </div>
        <div class="form-row">
          <label>
            <input type="checkbox" id="show-text" checked>
            <span data-i18n="showText">显示券码文字</span>
          </label>
        </div>
      </div>
    </div>

    <!-- 生成 PDF -->
    <button class="btn-generate" id="generate-btn" data-i18n="generate">生成 PDF</button>

  </aside>

  <!-- 右侧预览面板 -->
  <section class="preview-panel" id="preview-panel">
    <div class="preview-placeholder" id="preview-placeholder">
      <div class="ph-icon">🎨</div>
      <div class="ph-text" data-i18n="placeholder">上传模板图片后在此预览</div>
      <div class="ph-hint" data-i18n="placeholderHint">拖拽二维码和文字到目标位置</div>
    </div>

    <div class="preview-container hidden" id="preview-container">
      <img class="template-img" id="template-img" alt="template">

      <!-- 二维码元素 -->
      <div class="draggable qr-element" id="qr-element">
        <img id="qr-img" alt="qr">
        <div class="resize-handle resize-handle-qr" id="qr-resize"></div>
      </div>

      <!-- 券码文字元素 -->
      <div class="draggable text-element" id="text-element">
        <span id="text-content">COUPON001</span>
        <div class="resize-handle resize-handle-text" id="text-resize"></div>
      </div>

      <!-- 拖拽提示 -->
      <div class="drag-hint" id="drag-hint" data-i18n="dragHint">拖拽移动 · 拖角调整大小</div>
    </div>
  </section>
</div>

<!-- Toast 容器 -->
<div class="toast-container" id="toast-container"></div>

<!-- 加载遮罩 -->
<div class="loading-overlay" id="loading-overlay">
  <div class="spinner"></div>
  <div class="loading-text" id="loading-text" data-i18n="generating">正在生成 PDF，请稍候...</div>
</div>

<script>
// ==================== i18n 多语言 ====================
const I18N = {
  zh: {
    appTitle: '优惠券批量生成器',
    appSubtitle: '批量生成带独立二维码的优惠券 PDF',
    step1Title: '导入券码数据',
    step1Desc: '上传 Excel / CSV 文件，自动读取券码',
    uploadExcel: '选择 Excel 文件',
    uploadExcelHint: '支持 .xlsx .xls .et .csv',
    columnLabel: '券码所在列',
    skipHeader: '跳过第一行（标题行）',
    codesCount: '已读取',
    codesUnit: '个券码',
    clear: '清空',
    step2Title: '加载模板图片',
    step2Desc: '上传优惠券底图（含 logo、固定文案）',
    uploadImage: '选择模板图片',
    uploadImageHint: '支持 PNG / JPG / JPEG / BMP',
    imageSize: '图片尺寸：',
    step3Title: '排版定位',
    step3Desc: '拖拽调整二维码和文字位置，所见即所得',
    qrPosition: '二维码',
    textPosition: '券码文字',
    size: '大小',
    fontSize: '字号',
    showText: '显示券码文字',
    generate: '生成 PDF',
    generating: '正在生成 PDF，请稍候...',
    generateSuccess: 'PDF 生成成功！',
    generateError: '生成失败',
    noTemplate: '请先上传模板图片',
    noCodes: '请先导入券码数据',
    uploadSuccess: '上传成功',
    uploadError: '上传失败',
    clearSuccess: '已清空',
    parseError: '文件解析失败',
    invalidFile: '无效文件',
    dragHint: '拖拽移动 · 拖角调整大小',
    placeholder: '上传模板图片后在此预览',
    placeholderHint: '拖拽二维码和文字到目标位置',
  },
  en: {
    appTitle: 'Coupon Batch Generator',
    appSubtitle: 'Batch generate coupon PDFs with unique QR codes',
    step1Title: 'Import Coupon Codes',
    step1Desc: 'Upload Excel / CSV file to read codes',
    uploadExcel: 'Select Excel File',
    uploadExcelHint: 'Supports .xlsx .xls .et .csv',
    columnLabel: 'Code Column',
    skipHeader: 'Skip first row (header)',
    codesCount: 'Loaded',
    codesUnit: 'codes',
    clear: 'Clear',
    step2Title: 'Load Template Image',
    step2Desc: 'Upload coupon background image',
    uploadImage: 'Select Template Image',
    uploadImageHint: 'Supports PNG / JPG / JPEG / BMP',
    imageSize: 'Image size: ',
    step3Title: 'Position & Layout',
    step3Desc: 'Drag to adjust QR and text position',
    qrPosition: 'QR Code',
    textPosition: 'Coupon Text',
    size: 'Size',
    fontSize: 'Font',
    showText: 'Show coupon text',
    generate: 'Generate PDF',
    generating: 'Generating PDF, please wait...',
    generateSuccess: 'PDF generated successfully!',
    generateError: 'Generation failed',
    noTemplate: 'Please upload a template image first',
    noCodes: 'Please import coupon codes first',
    uploadSuccess: 'Upload successful',
    uploadError: 'Upload failed',
    clearSuccess: 'Cleared',
    parseError: 'File parsing failed',
    invalidFile: 'Invalid file',
    dragHint: 'Drag to move · Drag corner to resize',
    placeholder: 'Upload template image to preview here',
    placeholderHint: 'Drag QR code and text to position',
  }
};

// ==================== 状态 ====================
let currentLang = 'zh';
let templateLoaded = false;
let imgWidth = 0;
let imgHeight = 0;
let scale = 1;
let codesCount = 0;
let codesPreview = [];

// 配置（图像像素坐标）
let qrConfig = { x: 50, y: 50, size: 100 };
let textConfig = { x: 50, y: 200, fontSize: 24, show: true };

// ==================== DOM 引用 ====================
const $ = id => document.getElementById(id);
const excelInput = $('excel-input');
const columnSelect = $('column-select');
const skipHeader = $('skip-header');
const imageInput = $('image-input');
const previewContainer = $('preview-container');
const previewPlaceholder = $('preview-placeholder');
const templateImg = $('template-img');
const qrElement = $('qr-element');
const qrImg = $('qr-img');
const textElement = $('text-element');
const textContent = $('text-content');
const generateBtn = $('generate-btn');
const themeToggle = $('theme-toggle');
const langToggle = $('lang-toggle');
const dragHint = $('drag-hint');

// ==================== i18n 应用 ====================
function t(key) {
  return I18N[currentLang][key] || key;
}

function applyI18n() {
  document.querySelectorAll('[data-i18n]').forEach(el => {
    const key = el.getAttribute('data-i18n');
    if (I18N[currentLang][key]) {
      el.textContent = I18N[currentLang][key];
    }
  });
  document.documentElement.lang = currentLang === 'zh' ? 'zh' : 'en';
  langToggle.textContent = currentLang === 'zh' ? 'EN' : '中';
  updateCodesDisplay();
  updateImageInfo();
}

// ==================== 主题 ====================
function initTheme() {
  const saved = localStorage.getItem('coupon-theme') || 'light';
  setTheme(saved);
}

function setTheme(theme) {
  document.documentElement.setAttribute('data-theme', theme);
  localStorage.setItem('coupon-theme', theme);
  themeToggle.textContent = theme === 'dark' ? '☀️' : '🌙';
}

themeToggle.addEventListener('click', () => {
  const current = document.documentElement.getAttribute('data-theme');
  setTheme(current === 'dark' ? 'light' : 'dark');
});

langToggle.addEventListener('click', () => {
  currentLang = currentLang === 'zh' ? 'en' : 'zh';
  applyI18n();
});

// ==================== Toast 通知 ====================
function showToast(message, type = 'info') {
  const toast = document.createElement('div');
  toast.className = 'toast toast-' + type;
  toast.textContent = message;
  $('toast-container').appendChild(toast);
  requestAnimationFrame(() => toast.classList.add('show'));
  setTimeout(() => {
    toast.classList.remove('show');
    setTimeout(() => toast.remove(), 300);
  }, 3000);
}

// ==================== 加载遮罩 ====================
function showLoading(text) {
  $('loading-text').textContent = text || t('generating');
  $('loading-overlay').classList.add('show');
}

function hideLoading() {
  $('loading-overlay').classList.remove('show');
}

// ==================== 上传券码 ====================
async function uploadExcel() {
  const file = excelInput.files[0];
  if (!file) return;

  const col = columnSelect.value;
  const skip = skipHeader.checked ? 1 : 0;

  const formData = new FormData();
  formData.append('file', file);

  try {
    const resp = await fetch('/api/upload/excel?col=' + col + '&skip=' + skip + '&filename=' + encodeURIComponent(file.name), {
      method: 'POST',
      body: formData
    });
    const data = await resp.json();

    if (data.success) {
      codesCount = data.count;
      codesPreview = data.preview || [];
      $('excel-zone').classList.add('has-file');
      $('codes-info').classList.add('show');
      updateCodesDisplay();
      updateQRPreview();
      showToast(t('uploadSuccess') + ' (' + data.count + ')', 'success');
    } else {
      showToast(data.error || t('parseError'), 'error');
      excelInput.value = '';
    }
  } catch (err) {
    showToast(t('uploadError') + ': ' + err.message, 'error');
    excelInput.value = '';
  }
}

function updateCodesDisplay() {
  $('codes-count').textContent = codesCount;
  const previewEl = $('codes-preview');
  if (codesPreview.length > 0) {
    previewEl.classList.add('show');
    previewEl.innerHTML = codesPreview.map(c =>
      '<div class="code-item">' + escapeHtml(c) + '</div>'
    ).join('');
  } else {
    previewEl.classList.remove('show');
    previewEl.innerHTML = '';
  }
}

excelInput.addEventListener('change', uploadExcel);

// 列/跳过标题变化时重新上传（如果有文件）
let reuploadTimer;
function scheduleReupload() {
  clearTimeout(reuploadTimer);
  reuploadTimer = setTimeout(() => {
    if (excelInput.files[0]) uploadExcel();
  }, 100);
}
columnSelect.addEventListener('change', scheduleReupload);
skipHeader.addEventListener('change', scheduleReupload);

// ==================== 上传模板图片 ====================
async function uploadImage() {
  const file = imageInput.files[0];
  if (!file) return;

  const formData = new FormData();
  formData.append('file', file);

  try {
    const resp = await fetch('/api/upload/template', {
      method: 'POST',
      body: formData
    });
    const data = await resp.json();

    if (data.success) {
      templateLoaded = true;
      imgWidth = data.width;
      imgHeight = data.height;
      $('image-zone').classList.add('has-file');
      $('image-info').classList.add('show');
      updateImageInfo();
      loadTemplatePreview();
      showToast(t('uploadSuccess'), 'success');
    } else {
      showToast(data.error || t('uploadError'), 'error');
      imageInput.value = '';
    }
  } catch (err) {
    showToast(t('uploadError') + ': ' + err.message, 'error');
    imageInput.value = '';
  }
}

function updateImageInfo() {
  if (imgWidth > 0) {
    $('image-size-text').textContent = imgWidth + ' x ' + imgHeight + ' px';
  }
}

imageInput.addEventListener('change', uploadImage);

// ==================== 预览加载 ====================
function loadTemplatePreview() {
  previewPlaceholder.classList.add('hidden');
  previewContainer.classList.remove('hidden');

  templateImg.onload = function() {
    calculateScale();
    updateQRDisplay();
    updateTextDisplay();
    updateQRPreview();
  };
  templateImg.src = '/api/template?t=' + Date.now();
}

function calculateScale() {
  const panel = $('preview-panel');
  const maxW = panel.clientWidth - 48;
  const maxH = panel.clientHeight - 48;
  scale = Math.min(maxW / imgWidth, maxH / imgHeight, 1);
  if (scale === Infinity || isNaN(scale) || scale <= 0) scale = 1;

  const dispW = Math.round(imgWidth * scale);
  const dispH = Math.round(imgHeight * scale);
  templateImg.style.width = dispW + 'px';
  templateImg.style.height = dispH + 'px';
  previewContainer.style.width = dispW + 'px';
  previewContainer.style.height = dispH + 'px';
}

function updateQRPreview() {
  const code = (codesPreview && codesPreview[0]) ? codesPreview[0] : 'EXAMPLE';
  qrImg.src = '/api/qr?code=' + encodeURIComponent(code) + '&size=500&t=' + Date.now();
}

// ==================== 显示更新 ====================
function updateQRDisplay() {
  const dispX = qrConfig.x * scale;
  const dispY = qrConfig.y * scale;
  const dispSize = qrConfig.size * scale;
  qrElement.style.left = dispX + 'px';
  qrElement.style.top = dispY + 'px';
  qrElement.style.width = dispSize + 'px';
  qrElement.style.height = dispSize + 'px';
  $('qr-x').value = Math.round(qrConfig.x);
  $('qr-y').value = Math.round(qrConfig.y);
  $('qr-size').value = Math.round(qrConfig.size);
}

function updateTextDisplay() {
  const dispX = textConfig.x * scale;
  const dispY = textConfig.y * scale;
  const dispFontSize = textConfig.fontSize * scale;
  textElement.style.left = dispX + 'px';
  textElement.style.top = dispY + 'px';
  textElement.style.fontSize = dispFontSize + 'px';
  textElement.style.display = textConfig.show ? '' : 'none';
  textContent.textContent = (codesPreview && codesPreview[0]) ? codesPreview[0] : 'COUPON001';
  $('text-x').value = Math.round(textConfig.x);
  $('text-y').value = Math.round(textConfig.y);
  $('text-size').value = Math.round(textConfig.fontSize);
  $('show-text').checked = textConfig.show;
}

// ==================== 拖拽功能 ====================
function makeDraggable(element, config, type) {
  let isDragging = false;
  let startMouseX, startMouseY;
  let startConfigX, startConfigY;

  element.addEventListener('mousedown', function(e) {
    if (e.target.classList.contains('resize-handle')) return;
    isDragging = true;
    startMouseX = e.clientX;
    startMouseY = e.clientY;
    startConfigX = config.x;
    startConfigY = config.y;
    previewContainer.classList.add('dragging');
    element.classList.add('selected');
    showDragHint();
    e.preventDefault();
  });

  document.addEventListener('mousemove', function(e) {
    if (!isDragging) return;
    const dx = (e.clientX - startMouseX) / scale;
    const dy = (e.clientY - startMouseY) / scale;
    const elemSize = type === 'qr' ? config.size : config.fontSize;
    config.x = Math.max(0, Math.min(startConfigX + dx, imgWidth - elemSize));
    config.y = Math.max(0, Math.min(startConfigY + dy, imgHeight - elemSize));
    if (type === 'qr') {
      updateQRDisplay();
    } else {
      updateTextDisplay();
    }
  });

  document.addEventListener('mouseup', function() {
    if (isDragging) {
      isDragging = false;
      previewContainer.classList.remove('dragging');
      element.classList.remove('selected');
      hideDragHint();
    }
  });
}

// ==================== 缩放手柄 ====================
function makeResizable(handle, config, type) {
  let isResizing = false;
  let startMouseX, startMouseY;
  let startSize;

  handle.addEventListener('mousedown', function(e) {
    isResizing = true;
    startMouseX = e.clientX;
    startMouseY = e.clientY;
    startSize = type === 'qr' ? config.size : config.fontSize;
    previewContainer.classList.add('dragging');
    showDragHint();
    e.preventDefault();
    e.stopPropagation();
  });

  document.addEventListener('mousemove', function(e) {
    if (!isResizing) return;
    const dx = (e.clientX - startMouseX) / scale;
    const dy = (e.clientY - startMouseY) / scale;
    const delta = Math.max(dx, dy);
    let newSize = startSize + delta;

    if (type === 'qr') {
      newSize = Math.max(20, Math.min(500, newSize));
      config.size = newSize;
      // 保持位置不超出边界
      config.x = Math.min(config.x, imgWidth - newSize);
      config.y = Math.min(config.y, imgHeight - newSize);
      updateQRDisplay();
    } else {
      newSize = Math.max(8, Math.min(100, newSize));
      config.fontSize = newSize;
      config.x = Math.min(config.x, imgWidth - newSize * 3);
      config.y = Math.min(config.y, imgHeight - newSize);
      updateTextDisplay();
    }
  });

  document.addEventListener('mouseup', function() {
    if (isResizing) {
      isResizing = false;
      previewContainer.classList.remove('dragging');
      hideDragHint();
    }
  });
}

function showDragHint() {
  dragHint.classList.add('show');
}

function hideDragHint() {
  dragHint.classList.remove('show');
}

// ==================== 输入框联动 ====================
function bindInput(inputId, config, key, type) {
  $(inputId).addEventListener('input', function() {
    const val = parseInt(this.value) || 0;
    if (type === 'qr') {
      if (key === 'size') {
        config[key] = Math.max(20, Math.min(500, val));
      } else {
        config[key] = Math.max(0, Math.min(val, imgWidth - config.size));
      }
      updateQRDisplay();
    } else {
      if (key === 'fontSize') {
        config[key] = Math.max(8, Math.min(100, val));
      } else {
        config[key] = Math.max(0, val);
      }
      updateTextDisplay();
    }
  });
}

bindInput('qr-x', qrConfig, 'x', 'qr');
bindInput('qr-y', qrConfig, 'y', 'qr');
bindInput('qr-size', qrConfig, 'size', 'qr');
bindInput('text-x', textConfig, 'x', 'text');
bindInput('text-y', textConfig, 'y', 'text');
bindInput('text-size', textConfig, 'fontSize', 'text');

$('show-text').addEventListener('change', function() {
  textConfig.show = this.checked;
  updateTextDisplay();
});

// ==================== 清空操作 ====================
$('clear-excel').addEventListener('click', async function() {
  try {
    await fetch('/api/reset?what=excel', { method: 'POST' });
    codesCount = 0;
    codesPreview = [];
    excelInput.value = '';
    $('excel-zone').classList.remove('has-file');
    $('codes-info').classList.remove('show');
    $('codes-preview').classList.remove('show');
    $('codes-preview').innerHTML = '';
    updateQRPreview();
    showToast(t('clearSuccess'), 'info');
  } catch (err) {
    showToast(err.message, 'error');
  }
});

$('clear-image').addEventListener('click', async function() {
  try {
    await fetch('/api/reset?what=image', { method: 'POST' });
    templateLoaded = false;
    imgWidth = 0;
    imgHeight = 0;
    imageInput.value = '';
    $('image-zone').classList.remove('has-file');
    $('image-info').classList.remove('show');
    previewContainer.classList.add('hidden');
    previewPlaceholder.classList.remove('hidden');
    showToast(t('clearSuccess'), 'info');
  } catch (err) {
    showToast(err.message, 'error');
  }
});

// ==================== 生成 PDF ====================
generateBtn.addEventListener('click', async function() {
  if (!templateLoaded) {
    showToast(t('noTemplate'), 'error');
    return;
  }
  if (codesCount === 0) {
    showToast(t('noCodes'), 'error');
    return;
  }

  generateBtn.disabled = true;
  showLoading(t('generating'));

  try {
    const params = {
      qr_x: Math.round(qrConfig.x),
      qr_y: Math.round(qrConfig.y),
      qr_size: Math.round(qrConfig.size),
      text_x: Math.round(textConfig.x),
      text_y: Math.round(textConfig.y),
      text_font_size: Math.round(textConfig.fontSize),
      text_show: textConfig.show
    };

    const resp = await fetch('/api/generate', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify(params)
    });

    if (!resp.ok) {
      const err = await resp.json().catch(() => ({}));
      throw new Error(err.error || t('generateError'));
    }

    const blob = await resp.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = 'coupons.pdf';
    document.body.appendChild(a);
    a.click();
    a.remove();
    URL.revokeObjectURL(url);

    showToast(t('generateSuccess'), 'success');
  } catch (err) {
    showToast(err.message, 'error');
  } finally {
    generateBtn.disabled = false;
    hideLoading();
  }
});

// ==================== 工具函数 ====================
function escapeHtml(str) {
  const div = document.createElement('div');
  div.textContent = str;
  return div.innerHTML;
}

// ==================== 窗口大小变化 ====================
let resizeTimer;
window.addEventListener('resize', function() {
  clearTimeout(resizeTimer);
  resizeTimer = setTimeout(function() {
    if (templateLoaded) {
      calculateScale();
      updateQRDisplay();
      updateTextDisplay();
    }
  }, 150);
});

// ==================== 初始化 ====================
makeDraggable(qrElement, qrConfig, 'qr');
makeDraggable(textElement, textConfig, 'text');
makeResizable($('qr-resize'), qrConfig, 'qr');
makeResizable($('text-resize'), textConfig, 'text');

initTheme();
applyI18n();
</script>

</body>
</html>'''


# ==================== HTTP 请求处理器 ====================
class CouponHandler(http.server.BaseHTTPRequestHandler):

    def log_message(self, format, *args):
        """简化日志输出。"""
        sys.stderr.write("[%s] %s\n" % (self.log_date_time_string(), format % args))

    def send_json(self, data, status=200):
        """发送 JSON 响应。"""
        body = json.dumps(data, ensure_ascii=False).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_error_json(self, message, status=400):
        """发送错误 JSON。"""
        self.send_json({'success': False, 'error': message}, status)

    def do_GET(self):
        """处理 GET 请求。"""
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            params = parse_qs(parsed.query)

            if path == '/' or path == '/index.html':
                self._serve_html()
            elif path == '/api/qr':
                self._serve_qr(params)
            elif path == '/api/template':
                self._serve_template()
            else:
                self.send_error_json('Not found', 404)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.send_error_json(str(e), 500)

    def do_POST(self):
        """处理 POST 请求。"""
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            params = parse_qs(parsed.query)

            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length) if content_length > 0 else b''

            if path == '/api/upload/excel':
                self._handle_upload_excel(body, params)
            elif path == '/api/upload/template':
                self._handle_upload_template(body)
            elif path == '/api/generate':
                self._handle_generate(body)
            elif path == '/api/reset':
                self._handle_reset(params)
            else:
                self.send_error_json('Not found', 404)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.send_error_json(str(e), 500)

    # ---------- GET 处理器 ----------

    def _serve_html(self):
        body = HTML_TEMPLATE.encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _serve_qr(self, params):
        code = params.get('code', [''])[0]
        size_str = params.get('size', ['200'])[0]
        try:
            size = int(size_str)
            size = max(20, min(1000, size))
        except ValueError:
            size = 200

        if not code:
            self.send_error_json('Missing code parameter', 400)
            return

        img = generate_qr_image(code, size)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        data = buf.getvalue()

        self.send_response(200)
        self.send_header('Content-Type', 'image/png')
        self.send_header('Content-Length', str(len(data)))
        self.send_header('Cache-Control', 'no-cache')
        self.end_headers()
        self.wfile.write(data)

    def _serve_template(self):
        with state_lock:
            if not state['template_bytes']:
                self.send_error_json('No template uploaded', 404)
                return
            data = state['template_bytes']

        self.send_response(200)
        self.send_header('Content-Type', 'image/png')
        self.send_header('Content-Length', str(len(data)))
        self.send_header('Cache-Control', 'no-cache')
        self.end_headers()
        self.wfile.write(data)

    # ---------- POST 处理器 ----------

    def _handle_upload_excel(self, body, params):
        # 解析 multipart
        content_type = self.headers.get('Content-Type', '')
        if 'boundary=' not in content_type:
            self.send_error_json('Invalid content type', 400)
            return

        boundary = content_type.split('boundary=')[1].strip()
        files = parse_multipart(body, boundary)

        if 'file' not in files:
            self.send_error_json('No file uploaded', 400)
            return

        file_data = files['file']['content']
        filename = files['file']['filename'] or params.get('filename', ['unknown'])[0]

        # 保存到临时文件
        ext = os.path.splitext(filename)[1].lower()
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(file_data)
            tmp_path = tmp.name

        try:
            col = int(params.get('col', ['0'])[0])
            skip = params.get('skip', ['0'])[0] == '1'
            codes = parse_excel_file(tmp_path, col=col, skip_header=skip)

            with state_lock:
                state['codes'] = codes
                state['codes_filename'] = filename

            self.send_json({
                'success': True,
                'count': len(codes),
                'preview': codes[:10]
            })
        except ValueError as e:
            self.send_error_json(str(e), 400)
        except Exception as e:
            self.send_error_json(str(e), 400)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    def _handle_upload_template(self, body):
        content_type = self.headers.get('Content-Type', '')
        if 'boundary=' not in content_type:
            self.send_error_json('Invalid content type', 400)
            return

        boundary = content_type.split('boundary=')[1].strip()
        files = parse_multipart(body, boundary)

        if 'file' not in files:
            self.send_error_json('No file uploaded', 400)
            return

        file_data = files['file']['content']
        filename = files['file']['filename'] or 'template'

        try:
            img = Image.open(io.BytesIO(file_data))
            with state_lock:
                state['template_image'] = img
                state['template_size'] = img.size
                # 存储为 PNG bytes
                buf = io.BytesIO()
                img.save(buf, format='PNG')
                state['template_bytes'] = buf.getvalue()

            self.send_json({
                'success': True,
                'width': img.size[0],
                'height': img.size[1]
            })
        except Exception as e:
            self.send_error_json(str(e), 400)

    def _handle_generate(self, body):
        try:
            params = json.loads(body.decode('utf-8'))
        except (json.JSONDecodeError, UnicodeDecodeError):
            self.send_error_json('Invalid JSON body', 400)
            return

        with state_lock:
            if not state['codes']:
                self.send_error_json('No coupon codes loaded', 400)
                return
            if not state['template_image']:
                self.send_error_json('No template image loaded', 400)
                return

            codes = list(state['codes'])
            template_image = state['template_image'].copy()

        qr_config = {
            'x': int(params.get('qr_x', 50)),
            'y': int(params.get('qr_y', 50)),
            'size': int(params.get('qr_size', 100)),
        }
        text_config = {
            'x': int(params.get('text_x', 50)),
            'y': int(params.get('text_y', 200)),
            'font_size': int(params.get('text_font_size', 24)),
            'show': bool(params.get('text_show', True)),
        }

        try:
            pdf_buf = generate_pdf(codes, template_image, qr_config, text_config)
            pdf_data = pdf_buf.getvalue()
        except Exception as e:
            self.send_error_json(str(e), 500)
            return

        self.send_response(200)
        self.send_header('Content-Type', 'application/pdf')
        self.send_header('Content-Disposition', 'attachment; filename="coupons.pdf"')
        self.send_header('Content-Length', str(len(pdf_data)))
        self.end_headers()
        self.wfile.write(pdf_data)

    def _handle_reset(self, params):
        what = params.get('what', ['all'])[0]

        with state_lock:
            if what in ('excel', 'all'):
                state['codes'] = []
                state['codes_filename'] = None
            if what in ('image', 'all'):
                state['template_image'] = None
                state['template_bytes'] = None
                state['template_size'] = None

        self.send_json({'success': True, 'what': what})


# ==================== 示例文件生成 ====================
def create_example_files(output_dir='.'):
    """生成测试用示例文件。"""
    # 券码示例.xlsx
    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '券码'
        ws['A1'] = '券码'
        ws['B1'] = '面值'
        for i in range(1, 21):
            ws.cell(row=i + 1, column=1, value='COUPON%04d' % i)
            ws.cell(row=i + 1, column=2, value=100)
        xlsx_path = os.path.join(output_dir, '券码示例.xlsx')
        wb.save(xlsx_path)
        print('Created: ' + xlsx_path.encode('utf-8', errors='replace').decode('utf-8'))

    # 模板示例.png
    img = Image.new('RGB', (800, 500), (255, 252, 245))
    draw = ImageDraw.Draw(img)

    # 边框
    draw.rounded_rectangle([20, 20, 780, 480], radius=20, outline=(74, 158, 255), width=3)

    # 标题 — 跨平台字体检测
    font_path = _find_chinese_font()
    try:
        if font_path:
            font_large = ImageFont.truetype(font_path, 48)
            font_medium = ImageFont.truetype(font_path, 24)
            font_small = ImageFont.truetype(font_path, 16)
        else:
            font_large = ImageFont.load_default()
            font_medium = ImageFont.load_default()
            font_small = ImageFont.load_default()
    except (IOError, OSError):
        font_large = ImageFont.load_default()
        font_medium = ImageFont.load_default()
        font_small = ImageFont.load_default()

    draw.text((300, 60), '优 惠 券', fill=(26, 29, 40), font=font_large)
    draw.text((320, 130), 'COUPON', fill=(107, 114, 128), font=font_small)

    # 面值
    draw.text((280, 180), '¥ 100', fill=(255, 107, 107), font=font_large)

    # 说明文字
    draw.text((200, 280), '凭此券可抵扣 100 元', fill=(107, 114, 128), font=font_medium)
    draw.text((200, 320), '有效期至 2026-12-31', fill=(107, 114, 128), font=font_medium)
    draw.text((200, 360), '门店地址：广州市天河区', fill=(107, 114, 128), font=font_medium)

    # 虚线分割
    for x in range(60, 740, 10):
        draw.line([(x, 410), (x + 5, 410)], fill=(200, 200, 200), width=1)

    draw.text((250, 430), '请扫码核销', fill=(150, 150, 150), font=font_small)

    png_path = os.path.join(output_dir, '模板示例.png')
    img.save(png_path)
    print('Created: ' + png_path.encode('utf-8', errors='replace').decode('utf-8'))


# ==================== 主函数 ====================
def main():
    # 检查依赖
    missing = []
    if not HAS_OPENPYXL:
        missing.append('openpyxl')
    if not HAS_XLRD:
        missing.append('xlrd')
    if missing:
        print('Warning: Missing optional packages: ' + ', '.join(missing))
        print('Install with: pip install ' + ' '.join(missing))
        print()

    # 生成示例文件（如果不存在且带参数）
    if '--create-examples' in sys.argv:
        create_example_files()
        return

    # 启动服务器
    try:
        server = http.server.ThreadingHTTPServer((HOST, PORT), CouponHandler)
    except OSError as e:
        if 'Address already in use' in str(e):
            print('Port %d already in use, trying to open existing instance...' % PORT)
            webbrowser.open('http://%s:%d' % (HOST, PORT))
            return
        raise

    url = 'http://%s:%d' % (HOST, PORT)
    print('=' * 50)
    print('  优惠券批量生成器 / Coupon Batch Generator')
    print('=' * 50)
    print()
    print('  服务地址: ' + url)
    print('  按 Ctrl+C 停止服务')
    print()

    # 自动打开浏览器
    threading.Timer(0.5, lambda: webbrowser.open(url)).start()

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('\n正在关闭服务...')
        server.shutdown()
        server.server_close()
        print('服务已停止。')


if __name__ == '__main__':
    main()
