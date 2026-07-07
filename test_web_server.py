#!/usr/bin/env python3
"""
优惠券批量生成器 - 测试套件
============================
共 28 项测试，覆盖：
  - QR 码生成 (3)
  - Excel/CSV 解析 (10)
  - PDF 生成 (3)
  - HTTP 端点 (12)

运行: python3 test_web_server.py
"""

import unittest
import os
import sys
import io
import json
import http.client
import threading
import tempfile
import time

# 将当前目录加入路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import coupon_web


# ==================== 测试服务器辅助 ====================
class TestServer:
    """在随机端口启动测试服务器。"""

    def __init__(self, port=18888):
        self.port = port
        self.server = None
        self.thread = None

    def start(self):
        # 重置全局状态
        with coupon_web.state_lock:
            coupon_web.state['template_image'] = None
            coupon_web.state['template_bytes'] = None
            coupon_web.state['template_size'] = None
            coupon_web.state['codes'] = []
            coupon_web.state['codes_filename'] = None

        self.server = http.server.ThreadingHTTPServer(
            ('localhost', self.port), coupon_web.CouponHandler
        )
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)
        self.thread.start()
        time.sleep(0.2)

    def stop(self):
        if self.server:
            self.server.shutdown()
            self.server.server_close()

    def request(self, method, path, body=None, headers=None):
        conn = http.client.HTTPConnection('localhost', self.port, timeout=10)
        h = dict(headers) if headers else {}
        if body is not None and 'Content-Type' not in h:
            h['Content-Type'] = 'application/json'
        conn.request(method, path, body, h)
        resp = conn.getresponse()
        data = resp.read()
        conn.close()
        return resp.status, data, dict(resp.getheaders())


def make_multipart(fields, boundary='----TestBoundary12345'):
    """构造 multipart/form-data 请求体。"""
    lines = []
    for name, value in fields.items():
        if isinstance(value, dict):
            # 文件字段
            lines.append('--%s' % boundary)
            lines.append('Content-Disposition: form-data; name="%s"; filename="%s"' % (name, value['filename']))
            lines.append('Content-Type: %s' % value.get('content_type', 'application/octet-stream'))
            lines.append('')
            # 这里用 bytes
            header = '\r\n'.join(lines).encode('utf-8') + b'\r\n'
            return header + value['content'] + ('\r\n--%s--\r\n' % boundary).encode('utf-8')
        else:
            lines.append('--%s' % boundary)
            lines.append('Content-Disposition: form-data; name="%s"' % name)
            lines.append('')
            lines.append(str(value))
    body = '\r\n'.join(lines)
    return ('--%s--\r\n' % boundary).join(body.split('--%s--' % boundary)) if False else body.encode('utf-8') + ('\r\n--%s--\r\n' % boundary).encode('utf-8')


def make_multipart_body(filename, content, content_type='application/octet-stream', boundary='----TestBoundary12345'):
    """构造单个文件上传的 multipart body。"""
    header = (
        '--%s\r\n'
        'Content-Disposition: form-data; name="file"; filename="%s"\r\n'
        'Content-Type: %s\r\n\r\n'
    ) % (boundary, filename, content_type)
    return header.encode('utf-8') + content + ('\r\n--%s--\r\n' % boundary).encode('utf-8')


# ==================== QR 码生成测试 (3) ====================
class TestQRGeneration(unittest.TestCase):
    """测试 QR 码生成功能。"""

    def test_01_qr_generates_correct_size(self):
        """QR 码生成指定尺寸的图片。"""
        img = coupon_web.generate_qr_image('TEST123', 200)
        self.assertEqual(img.size, (200, 200))
        self.assertEqual(img.mode, 'RGB')

    def test_02_qr_different_codes_produce_different_images(self):
        """不同券码生成不同的二维码。"""
        img1 = coupon_web.generate_qr_image('CODE1', 100)
        img2 = coupon_web.generate_qr_image('CODE2', 100)
        self.assertNotEqual(list(img1.getdata()), list(img2.getdata()))

    def test_03_qr_various_sizes(self):
        """各种尺寸都能正确生成。"""
        for size in [20, 50, 100, 200, 500]:
            img = coupon_web.generate_qr_image('TEST', size)
            self.assertEqual(img.size, (size, size))


# ==================== Excel/CSV 解析测试 (10) ====================
class TestExcelParsing(unittest.TestCase):
    """测试 Excel/CSV 文件解析。"""

    def setUp(self):
        self.tempdir = tempfile.mkdtemp()

    def test_04_csv_basic(self):
        """CSV 基本读取。"""
        filepath = os.path.join(self.tempdir, 'test.csv')
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('COUPON001\nCOUPON002\nCOUPON003\n')
        codes = coupon_web.parse_excel_file(filepath, col=0, skip_header=False)
        self.assertEqual(len(codes), 3)
        self.assertEqual(codes[0], 'COUPON001')
        self.assertEqual(codes[2], 'COUPON003')

    def test_05_csv_with_header_skip(self):
        """CSV 跳过标题行。"""
        filepath = os.path.join(self.tempdir, 'test.csv')
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('券码\nCOUPON001\nCOUPON002\n')
        codes = coupon_web.parse_excel_file(filepath, col=0, skip_header=True)
        self.assertEqual(len(codes), 2)
        self.assertEqual(codes[0], 'COUPON001')

    def test_06_csv_column_b(self):
        """CSV 读取 B 列。"""
        filepath = os.path.join(self.tempdir, 'test.csv')
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('A1,B1\nA2,COUPON001\nA3,COUPON002\n')
        codes = coupon_web.parse_excel_file(filepath, col=1, skip_header=True)
        self.assertEqual(len(codes), 2)
        self.assertEqual(codes[0], 'COUPON001')

    def test_07_csv_skip_empty_cells(self):
        """CSV 跳过空单元格。"""
        filepath = os.path.join(self.tempdir, 'test.csv')
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('CODE1\n\nCODE3\n')
        codes = coupon_web.parse_excel_file(filepath, col=0, skip_header=False)
        self.assertEqual(len(codes), 2)
        self.assertEqual(codes[0], 'CODE1')
        self.assertEqual(codes[1], 'CODE3')

    def test_08_csv_utf8_bom(self):
        """CSV UTF-8 BOM 编码兼容。"""
        filepath = os.path.join(self.tempdir, 'test.csv')
        with open(filepath, 'w', encoding='utf-8-sig') as f:
            f.write('券码\nCODE1\nCODE2\n')
        codes = coupon_web.parse_excel_file(filepath, col=0, skip_header=True)
        self.assertEqual(len(codes), 2)

    def test_09_xlsx_basic(self):
        """XLSX 基本读取。"""
        if not coupon_web.HAS_OPENPYXL:
            self.skipTest("openpyxl not installed")
        import openpyxl
        filepath = os.path.join(self.tempdir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '券码'
        for i in range(1, 6):
            ws.cell(row=i + 1, column=1, value='COUPON%03d' % i)
        wb.save(filepath)
        codes = coupon_web.parse_excel_file(filepath, col=0, skip_header=True)
        self.assertEqual(len(codes), 5)
        self.assertEqual(codes[0], 'COUPON001')

    def test_10_xlsx_column_c(self):
        """XLSX 读取 C 列。"""
        if not coupon_web.HAS_OPENPYXL:
            self.skipTest("openpyxl not installed")
        import openpyxl
        filepath = os.path.join(self.tempdir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Name'
        ws['C1'] = 'Code'
        ws['C2'] = 'CODE001'
        ws['C3'] = 'CODE002'
        wb.save(filepath)
        codes = coupon_web.parse_excel_file(filepath, col=2, skip_header=True)
        self.assertEqual(len(codes), 2)
        self.assertEqual(codes[0], 'CODE001')

    def test_11_xlsx_no_skip_header(self):
        """XLSX 不跳过标题行。"""
        if not coupon_web.HAS_OPENPYXL:
            self.skipTest("openpyxl not installed")
        import openpyxl
        filepath = os.path.join(self.tempdir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'HEADER'
        ws['A2'] = 'CODE001'
        ws['A3'] = 'CODE002'
        wb.save(filepath)
        codes = coupon_web.parse_excel_file(filepath, col=0, skip_header=False)
        self.assertEqual(len(codes), 3)

    def test_12_unsupported_format(self):
        """不支持的文件格式抛出异常。"""
        filepath = os.path.join(self.tempdir, 'test.txt')
        with open(filepath, 'w') as f:
            f.write('test')
        with self.assertRaises(ValueError):
            coupon_web.parse_excel_file(filepath, col=0)

    def test_13_empty_file(self):
        """空文件返回空列表。"""
        filepath = os.path.join(self.tempdir, 'empty.csv')
        with open(filepath, 'w') as f:
            f.write('')
        codes = coupon_web.parse_excel_file(filepath, col=0, skip_header=False)
        self.assertEqual(len(codes), 0)


# ==================== PDF 生成测试 (3) ====================
class TestPDFGeneration(unittest.TestCase):
    """测试 PDF 生成功能。"""

    def setUp(self):
        from PIL import Image
        self.template = Image.new('RGB', (400, 300), (255, 255, 255))

    def test_14_pdf_basic_generation(self):
        """PDF 基本生成。"""
        codes = ['CODE001', 'CODE002', 'CODE003']
        qr_config = {'x': 50, 'y': 50, 'size': 100}
        text_config = {'x': 50, 'y': 200, 'font_size': 24, 'show': True}
        buf = coupon_web.generate_pdf(codes, self.template, qr_config, text_config)
        data = buf.getvalue()
        self.assertGreater(len(data), 0)
        # PDF 文件头
        self.assertTrue(data.startswith(b'%PDF'))

    def test_15_pdf_multiple_codes(self):
        """PDF 生成多个券码页面。"""
        codes = ['CODE%03d' % i for i in range(1, 11)]
        qr_config = {'x': 50, 'y': 50, 'size': 80}
        text_config = {'x': 50, 'y': 200, 'font_size': 20, 'show': True}
        buf = coupon_web.generate_pdf(codes, self.template, qr_config, text_config)
        data = buf.getvalue()
        self.assertGreater(len(data), 1000)  # 至少有内容
        # 检查页数（通过 Page 对象计数）
        page_count = data.count(b'/Type /Page')
        # /Type /Page 可能出现在 Pages 和 Page 中，至少应有10
        self.assertGreaterEqual(page_count, 10)

    def test_16_pdf_text_hidden(self):
        """PDF 文字隐藏时仍能生成。"""
        codes = ['CODE001']
        qr_config = {'x': 50, 'y': 50, 'size': 100}
        text_config = {'x': 50, 'y': 200, 'font_size': 24, 'show': False}
        buf = coupon_web.generate_pdf(codes, self.template, qr_config, text_config)
        data = buf.getvalue()
        self.assertGreater(len(data), 0)
        self.assertTrue(data.startswith(b'%PDF'))

    def test_17_pdf_rgba_template(self):
        """RGBA 模板图片正确处理。"""
        from PIL import Image
        template = Image.new('RGBA', (300, 200), (255, 0, 0, 128))
        codes = ['CODE001']
        qr_config = {'x': 20, 'y': 20, 'size': 60}
        text_config = {'x': 20, 'y': 100, 'font_size': 16, 'show': True}
        buf = coupon_web.generate_pdf(codes, template, qr_config, text_config)
        data = buf.getvalue()
        self.assertTrue(data.startswith(b'%PDF'))


# ==================== HTTP 端点测试 (12) ====================
class TestHTTPServer(unittest.TestCase):
    """测试 HTTP 服务器端点。"""

    @classmethod
    def setUpClass(cls):
        cls.server = TestServer(port=18888)
        cls.server.start()
        # 预上传测试数据
        cls._setup_test_data()

    @classmethod
    def _setup_test_data(cls):
        """上传 CSV 和模板图片，供后续测试使用。"""
        # 上传 CSV
        csv_content = '券码\nCODE001\nCODE002\nCODE003\n'.encode('utf-8')
        body = make_multipart_body('test.csv', csv_content, 'text/csv')
        cls.server.request('POST', '/api/upload/excel?col=0&skip=1&filename=test.csv', body, {
            'Content-Type': 'multipart/form-data; boundary=----TestBoundary12345'
        })

        # 上传模板图片
        from PIL import Image
        img = Image.new('RGB', (400, 300), (255, 200, 100))
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        body = make_multipart_body('template.png', buf.getvalue(), 'image/png')
        cls.server.request('POST', '/api/upload/template', body, {
            'Content-Type': 'multipart/form-data; boundary=----TestBoundary12345'
        })

    @classmethod
    def tearDownClass(cls):
        cls.server.stop()

    def test_18_get_homepage(self):
        """GET / 返回 HTML 页面。"""
        status, data, _ = self.server.request('GET', '/')
        self.assertEqual(status, 200)
        self.assertIn(b'<html', data.lower())
        self.assertIn('优惠券'.encode('utf-8'), data)

    def test_19_get_qr_code(self):
        """GET /api/qr 返回 PNG 图片。"""
        status, data, _ = self.server.request('GET', '/api/qr?code=TEST&size=100')
        self.assertEqual(status, 200)
        self.assertTrue(data.startswith(b'\x89PNG'))

    def test_20_get_qr_default_size(self):
        """GET /api/qr 默认尺寸。"""
        status, data, _ = self.server.request('GET', '/api/qr?code=TEST')
        self.assertEqual(status, 200)
        self.assertTrue(data.startswith(b'\x89PNG'))

    def test_21_get_template_no_upload_then_404(self):
        """GET /api/template 未上传时返回 404。"""
        # 先重置
        self.server.request('POST', '/api/reset?what=image')
        status, _, _ = self.server.request('GET', '/api/template')
        self.assertEqual(status, 404)
        # 重新上传
        from PIL import Image
        img = Image.new('RGB', (200, 150), (100, 200, 50))
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        body = make_multipart_body('template.png', buf.getvalue(), 'image/png')
        self.server.request('POST', '/api/upload/template', body, {
            'Content-Type': 'multipart/form-data; boundary=----TestBoundary12345'
        })

    def test_22_upload_excel_csv(self):
        """POST /api/upload/excel 上传 CSV 文件。"""
        csv_content = '券码\nCODE001\nCODE002\nCODE003\n'.encode('utf-8')
        body = make_multipart_body('test.csv', csv_content, 'text/csv')
        status, data, _ = self.server.request('POST', '/api/upload/excel?col=0&skip=1&filename=test.csv', body, {
            'Content-Type': 'multipart/form-data; boundary=----TestBoundary12345'
        })
        self.assertEqual(status, 200)
        result = json.loads(data)
        self.assertTrue(result['success'])
        self.assertEqual(result['count'], 3)
        self.assertEqual(result['preview'][0], 'CODE001')

    def test_23_upload_excel_invalid_format(self):
        """POST /api/upload/excel 上传无效格式返回错误。"""
        body = make_multipart_body('test.txt', b'invalid content', 'text/plain')
        status, data, _ = self.server.request('POST', '/api/upload/excel?col=0&skip=0&filename=test.txt', body, {
            'Content-Type': 'multipart/form-data; boundary=----TestBoundary12345'
        })
        self.assertEqual(status, 400)

    def test_24_upload_template_image(self):
        """POST /api/upload/template 上传模板图片。"""
        from PIL import Image
        img = Image.new('RGB', (500, 350), (200, 100, 50))
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        body = make_multipart_body('template.png', buf.getvalue(), 'image/png')
        status, data, _ = self.server.request('POST', '/api/upload/template', body, {
            'Content-Type': 'multipart/form-data; boundary=----TestBoundary12345'
        })
        self.assertEqual(status, 200)
        result = json.loads(data)
        self.assertTrue(result['success'])
        self.assertEqual(result['width'], 500)
        self.assertEqual(result['height'], 350)

    def test_25_generate_pdf_success(self):
        """POST /api/generate 成功生成 PDF。"""
        # 确保有券码和模板
        csv_content = '券码\nCODE001\nCODE002\n'.encode('utf-8')
        body = make_multipart_body('test.csv', csv_content, 'text/csv')
        self.server.request('POST', '/api/upload/excel?col=0&skip=1&filename=test.csv', body, {
            'Content-Type': 'multipart/form-data; boundary=----TestBoundary12345'
        })

        params = json.dumps({
            'qr_x': 50, 'qr_y': 50, 'qr_size': 100,
            'text_x': 50, 'text_y': 200, 'text_font_size': 24, 'text_show': True
        })
        status, data, _ = self.server.request('POST', '/api/generate', params.encode('utf-8'))
        self.assertEqual(status, 200)
        self.assertTrue(data.startswith(b'%PDF'))

    def test_26_generate_no_codes_error(self):
        """POST /api/generate 无券码时返回 400。"""
        self.server.request('POST', '/api/reset?what=excel')
        params = json.dumps({
            'qr_x': 50, 'qr_y': 50, 'qr_size': 100,
            'text_x': 50, 'text_y': 200, 'text_font_size': 24, 'text_show': True
        })
        status, data, _ = self.server.request('POST', '/api/generate', params.encode('utf-8'))
        self.assertEqual(status, 400)
        # 恢复数据
        csv_content = '券码\nCODE001\nCODE002\n'.encode('utf-8')
        body = make_multipart_body('test.csv', csv_content, 'text/csv')
        self.server.request('POST', '/api/upload/excel?col=0&skip=1&filename=test.csv', body, {
            'Content-Type': 'multipart/form-data; boundary=----TestBoundary12345'
        })

    def test_27_reset_excel(self):
        """POST /api/reset?what=excel 清空券码。"""
        status, data, _ = self.server.request('POST', '/api/reset?what=excel')
        self.assertEqual(status, 200)
        result = json.loads(data)
        self.assertTrue(result['success'])
        self.assertEqual(result['what'], 'excel')
        # 恢复数据
        csv_content = '券码\nCODE001\n'.encode('utf-8')
        body = make_multipart_body('test.csv', csv_content, 'text/csv')
        self.server.request('POST', '/api/upload/excel?col=0&skip=1&filename=test.csv', body, {
            'Content-Type': 'multipart/form-data; boundary=----TestBoundary12345'
        })

    def test_28_reset_all(self):
        """POST /api/reset?what=all 清空所有。"""
        status, data, _ = self.server.request('POST', '/api/reset?what=all')
        self.assertEqual(status, 200)
        result = json.loads(data)
        self.assertTrue(result['success'])
        self.assertEqual(result['what'], 'all')


# ==================== 运行测试 ====================
if __name__ == '__main__':
    print('=' * 60)
    print('  优惠券批量生成器 - 测试套件 (28 tests)')
    print('  Coupon Batch Generator - Test Suite')
    print('=' * 60)
    print()

    unittest.main(verbosity=2)
